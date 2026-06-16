from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine

from app.db import get_session
import app.main as main_module
import app.services.search_runner as search_runner


class EmptyCrawlerConnector:
    def __init__(self):
        self.last_status = {}

    def search(self, intent):
        self.last_status = {
            platform.value: {
                "query": f"site:{platform.value}.com skincare creator contact",
                "returned_count": 0,
                "parsed_count": 0,
                "fetched_count": 0,
                "skipped_count": 0,
                "errors": [],
            }
            for platform in intent.platforms
        }
        return []


class NoLiveSettings:
    youtube_api_key = ""
    tavily_api_key = ""
    search_engine_api_key = ""
    search_engine_id = ""


@pytest.fixture
def session_engine(tmp_path):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'test.db'}",
        connect_args={"check_same_thread": False},
    )
    SQLModel.metadata.create_all(engine)
    return engine


@pytest.fixture
def client(session_engine, monkeypatch):
    def override_session():
        with Session(session_engine) as session:
            yield session

    monkeypatch.setattr(main_module, "init_db", lambda: None)
    previous_overrides = dict(main_module.app.dependency_overrides)
    main_module.app.dependency_overrides[get_session] = override_session
    with TestClient(main_module.app, raise_server_exceptions=False) as test_client:
        yield test_client
    main_module.app.dependency_overrides = previous_overrides


def test_acceptance_search_to_results_to_export_and_card(client, monkeypatch):
    monkeypatch.setattr(search_runner, "get_settings", lambda: NoLiveSettings())
    monkeypatch.setattr(search_runner, "WebCrawlerConnector", EmptyCrawlerConnector)

    create_response = client.post(
        "/search",
        data={
            "input_text": "skincare",
            "platforms": ["youtube", "tiktok", "instagram"],
            "run_inline": "yes",
            "use_demo_data": "yes",
        },
        follow_redirects=False,
    )
    assert create_response.status_code in {302, 303}
    task_url = create_response.headers["location"]

    detail_response = client.get(task_url)
    assert detail_response.status_code == 200
    assert "搜索任务" in detail_response.text
    assert "公开爬虫搜索" in detail_response.text
    assert "演示数据" in detail_response.text
    assert "Creator A" in detail_response.text
    assert "Creator B" in detail_response.text
    assert "Creator C" in detail_response.text
    assert detail_response.text.count("审核卡片") >= 24
    assert "/export.xlsx" in detail_response.text
    assert "/results/" in detail_response.text
    assert "/card" in detail_response.text

    export_response = client.get(f"{task_url}/export.xlsx")
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export_response.content))
    sheet = workbook.active
    assert sheet["A1"].value == "达人"
    assert sheet.max_row >= 25
    exported_names = {sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)}
    exported_platforms = {sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)}
    assert {"Creator A", "Creator B", "Creator C"}.issubset(exported_names)
    assert exported_platforms == {"youtube", "tiktok", "instagram"}

    card_path = _extract_first_card_path(detail_response.text)
    card_response = client.get(card_path)
    assert card_response.status_code == 200
    assert any(name in card_response.text for name in {"Creator A", "Creator B", "Creator C"})
    assert "推荐理由：" in card_response.text


def _extract_first_card_path(html: str) -> str:
    marker = 'href="/results/'
    start = html.index(marker) + len('href="')
    end = html.index('"', start)
    return html[start:end]
