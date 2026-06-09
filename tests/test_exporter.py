from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine

from app.db import get_session
import app.main as main_module
from app.models import (
    ContactRecord,
    ContentSample,
    Creator,
    Platform,
    PlatformAccount,
    ScoreResult,
    SearchTask,
    TaskStatus,
)
from app.services.exporter import build_candidate_workbook


@pytest.fixture
def session_engine(tmp_path):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'test.db'}",
        connect_args={"check_same_thread": False},
    )
    SQLModel.metadata.create_all(engine)
    return engine


@pytest.fixture
def client(session_engine, monkeypatch):
    def override_session():
        with Session(session_engine) as session:
            yield session

    monkeypatch.setattr(main_module, "init_db", lambda: None)
    previous_overrides = dict(main_module.app.dependency_overrides)
    main_module.app.dependency_overrides[get_session] = override_session
    with TestClient(main_module.app, raise_server_exceptions=False) as test_client:
        yield test_client
    main_module.app.dependency_overrides = previous_overrides


@pytest.fixture
def db_session(session_engine):
    with Session(session_engine) as session:
        yield session


def test_build_candidate_workbook_contains_expected_headers():
    data = [
        {
            "creator": "Creator A",
            "platform": "YouTube",
            "followers": 310000,
            "recent_views": 96000,
            "engagement_rate": 0.071,
            "score": 87.0,
            "contact": "business@example.com",
        }
    ]
    payload = build_candidate_workbook(data)
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1][:7]] == [
        "Creator",
        "Platform",
        "Followers",
        "Recent Views",
        "Engagement Rate",
        "Score",
        "Contact",
    ]
    assert sheet["G2"].value == "business@example.com"


def test_build_candidate_workbook_escapes_formula_like_text_values():
    data = [
        {
            "creator": "=HYPERLINK(\"https://example.com\")",
            "platform": "youtube",
            "followers": 310000,
            "recent_views": 96000,
            "engagement_rate": 0.071,
            "score": 87.0,
            "contact": "@attacker",
        },
        {
            "creator": "+SUM(1,1)",
            "platform": "tiktok",
            "followers": 420000,
            "recent_views": 185000,
            "engagement_rate": 0.08,
            "score": 91.0,
            "contact": "-cmd",
        },
    ]

    payload = build_candidate_workbook(data)
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook.active

    assert sheet["A2"].value == "'=HYPERLINK(\"https://example.com\")"
    assert sheet["G2"].value == "'@attacker"
    assert sheet["A3"].value == "'+SUM(1,1)"
    assert sheet["G3"].value == "'-cmd"


def test_export_task_returns_candidate_workbook_with_persisted_metrics(client, db_session):
    creator = Creator(display_name="Creator A", primary_topics="skincare")
    account = PlatformAccount(
        creator=creator,
        platform=Platform.youtube,
        handle="@creator",
        profile_url="https://example.com/creator",
        follower_count=310000,
    )
    task = SearchTask(
        input_text="skincare",
        input_type="keyword",
        platforms="youtube",
        status=TaskStatus.complete,
    )
    db_session.add(account)
    db_session.add(task)
    db_session.flush()
    db_session.add(
        ContactRecord(
            creator_id=creator.id,
            contact_type="email",
            value="business@example.com",
            source_url="https://example.com/creator",
        )
    )
    db_session.add(
        ContentSample(
            account_id=account.id,
            content_url="https://example.com/video-1",
            view_count=1000,
            like_count=80,
            comment_count=20,
            share_count=10,
        )
    )
    db_session.add(
        ContentSample(
            account_id=account.id,
            content_url="https://example.com/video-2",
            view_count=2000,
            like_count=160,
            comment_count=40,
            share_count=20,
        )
    )
    db_session.add(
        ScoreResult(
            task_id=task.id,
            creator_id=creator.id,
            platform_account_id=account.id,
            normalized_engagement=0.071,
            final_score=87.0,
        )
    )
    db_session.commit()

    response = client.get(f"/tasks/{task.id}/export.xlsx")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response.headers["content-disposition"]
        == f'attachment; filename="task-{task.id}-candidates.xlsx"'
    )
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet["A2"].value == "Creator A"
    assert sheet["B2"].value == "youtube"
    assert sheet["D2"].value == 1500
    assert sheet["E2"].value == 0.11
    assert sheet["F2"].value == 87.0
    assert sheet["G2"].value == "business@example.com"


def test_export_task_returns_404_for_missing_task(client):
    response = client.get("/tasks/999/export.xlsx")

    assert response.status_code == 404
